import streamlit as st
from pandas import read_excel
from openpyxl import load_workbook

st.set_page_config(
    page_title="Fake generator", 
    page_icon=None, 
    layout="wide")

st.title("""Fake chart""")

flag = 0

loading = st.progress(0)

HUBS = ['AlleppeyHub_ALP','CharumoodHub_COO','ChengannurHub_CNN', 'Cherthala_CTA',
            'KaruvattaHub_KVT', 'KayamkulamHub_KKM', 'SASThuravoorODH_THO', 'STCMuthukulamODH_MKL']

uploaded_file = st.file_uploader("""Upload fake raw data file: """)


def table_styler(dataframe):
    return dataframe.style.set_table_styles(
        [{'selector' : 'th',
            'props' : [('background', '#ea6154'),
                        ('color' , '#fff3ef'),
                        ('font-family', 'helvetica'),
                        ('text-align', 'left')]},
        {'selector' : 'td',
            'props' : [('text-align', 'left'),
                        ('font-family', 'helvetica')]
        },
        {'selector' : 'tr:nth-of-type(odd)',
            'props' : [('background', '#DCDCDC')]
        },
        {'selector' : 'tr:nth-of-type(even)',
            'props' : [('background', 'white')]
        }
        ]
    )

if uploaded_file == None:
    
    st.write("""
    ### Load file using browse files!
    """)
else:
    if 'uploaded_file' not in st.session_state:
        st.session_state['uploaded_file'] = uploaded_file
    if st.session_state['uploaded_file'] != uploaded_file:
        st.session_state['uploaded_file'] = uploaded_file
        del st.session_state['filtered_df']
        
    st.write(""" ## Alleppey cluster Fake attempts""")
    loading.progress(10)
    my_table = st.table()

    if 'filtered_df' in st.session_state:
        filtered_df = st.session_state['filtered_df']
        flag = 2
        loading.progress(30)
    else:
        loading.progress(20)
        possible_sheet_name = ["raw data", "fake tids"]

        wb = load_workbook(uploaded_file)
        
        sheetname = [book for book in wb.sheetnames if (book.lower() in possible_sheet_name)]

        try:
            df = read_excel(uploaded_file, sheet_name=sheetname[0])
            flag = 1
            loading.progress(30)
        except Exception as e:
            st.error(e)



if (flag):
    loading.progress(40)
    if flag != 2:
        if 'Hub Name' in df:
            st.session_state['hub_name'] = hub_name = 'Hub Name'
        elif 'hub_name' in df:
            st.session_state['hub_name'] = hub_name = 'hub_name'
        # elif '' in df:
            # hub_name = ''

        COLUMNS = [hub_name, 'fake_detection_reason', 'geo_distance' ,'vendor_tracking_id', 'undel_unpick_status',
                    'agent_name', 'Kirana']
        
        EXCLUSION_LIST = ['DELIVERED', 'UNDELIVERED']
        loading.progress(60)
        filtered_df = df.loc[df[hub_name].isin(HUBS), COLUMNS].sort_values([hub_name,'Kirana'])
        filtered_df = filtered_df.loc[~filtered_df['undel_unpick_status'].isin(EXCLUSION_LIST)]
        filtered_df.loc[:, "geo_distance"] = filtered_df["geo_distance"].map('{:.2f}'.format)
        loading.progress(80)
        

    if 'filtered_df' not in st.session_state:
        st.session_state['filtered_df'] = filtered_df

    my_table.table(table_styler(filtered_df))
    loading.progress(100)
    hub_name = st.session_state['hub_name']
    my_form = st.form(key="filter_table")
    selected = my_form.multiselect("Choose hubs", HUBS)
    selected_df = filtered_df.loc[filtered_df[hub_name].isin(selected)].sort_values([hub_name,'Kirana'])
    update_table =my_form.form_submit_button()

    if update_table:
        styled_df = table_styler(selected_df)
        my_table.table(styled_df)
    loading.empty()